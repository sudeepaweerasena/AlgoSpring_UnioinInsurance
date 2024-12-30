import openpyxl
import pandas as pd
from datetime import datetime

# File paths and sheet names
OUTPUT_FILE_PATH = "D:\\AlgoSpring\\python\\Union_Insurance\\Standard Plans Template.xlsx"
OUTPUT_SHEET_NAME = "Sheet1"
INPUT_FILE_PATH = "D:\\AlgoSpring\\python\\Union_Insurance\\FromFile.xlsx"
INPUT_SHEET_NAME = "Sheet1"

def nlg_transfer_medical_data():
    try:
        input_df = pd.read_excel(INPUT_FILE_PATH, sheet_name=INPUT_SHEET_NAME)
        print("Initial Input Data:")
        print(input_df.head())

    except FileNotFoundError:
        print(f"Error: The file {INPUT_FILE_PATH} does not exist.")
        return
    except ValueError:
        print(f"Error: The sheet {INPUT_SHEET_NAME} does not exist in {INPUT_FILE_PATH}.")
        return

    # Replace 'Principal' with 'Employee' in the 'Relation' column
    input_df['Relation'] = input_df['Relation'].replace('Principal', 'Employee')
    print("After Replacing 'Principal' with 'Employee':")
    print(input_df.head())

    # Convert Gender to 'M' or 'F'
    input_df['Gender'] = input_df['Gender'].apply(lambda gender: 'M' if gender == 'Male' else ('F' if gender == 'Female' else gender))
    print("After Converting Gender:")
    print(input_df.head())

    # Convert Salary Type to 'Yes' or 'No'
    input_df['Salary Type'] = input_df['Salary Type'].apply(lambda salary_type: 'Yes' if salary_type == 'LSB' else ('No' if salary_type == 'HSB' else salary_type))
    print("After Converting Salary Type:")
    print(input_df.head())

    # Rename columns as per template needs
    column_mapping = {
        "Relation": "Relation",
        "Gender": "Gender",
        "DOB": "DOB",
        "Category": "Category",
        "Marital status": "Marital Status",
        "Visa Issued Emirates": "Emirate",
        "Salary Type": "LSB"
    }
    input_df.rename(columns=column_mapping, inplace=True)

    # Load the output workbook and sheet
    try:
        output_wb = openpyxl.load_workbook(OUTPUT_FILE_PATH)
        output_ws = output_wb[OUTPUT_SHEET_NAME]
    except FileNotFoundError:
        print(f"Error: The file {OUTPUT_FILE_PATH} does not exist.")
        return
    except KeyError:
        print(f"Error: The sheet {OUTPUT_SHEET_NAME} does not exist in {OUTPUT_FILE_PATH}.")
        return

    # Apply the DOB formatting
    for index, row in input_df.iterrows():

        output_ws.cell(row=index + 2, column=2).value = row["DOB"]
        output_ws.cell(row=index + 2, column=1).value = index + 1
        output_ws.cell(row=index + 2, column=5).value = row["Marital Status"]
        output_ws.cell(row=index + 2, column=3).value = row["Gender"]
        output_ws.cell(row=index + 2, column=4).value = row["Relation"]
        output_ws.cell(row=index + 2, column=7).value = row["LSB"]
        output_ws.cell(row=index + 2, column=8).value = row["Emirate"]
        category_letter = row["Category"]
        category_mapping = {'A': "Category A", 'B': "Category B", 'C': "Category C"}
        category_value = category_mapping.get(category_letter, "Unknown")
        output_ws.cell(row=index + 2, column=6).value = category_value

    output_wb.save(OUTPUT_FILE_PATH)
    print(f"Data successfully written to {OUTPUT_FILE_PATH}")

# # Call the function to execute
# nlg_transfer_medical_data()
